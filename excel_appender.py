#import openpyxl
from openpyxl import load_workbook
#https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#from openpyxl.styles import colors
from openpyxl.styles import Font, Color
#from openpyxl import Workbook
from openpyxl.utils  import get_column_letter
from copy import copy

class appender():
    def __init__(self,
                 filename1,
                 filename2,
                 H1,
                 sheet_name = None,
                 offset = 4
                 ):
        wb1 = load_workbook(filename1)
        wb2 = load_workbook(filename2)
        if sheet_name is None: ## if sheet name is not defined, the app is default to apply format function to all common sheets of two excel files
            sheet_name1 = wb1.sheetnames
            sheet_name2 = wb2.sheetnames
            sheet_name = list(set(sheet_name1).intersection(set(sheet_name2)))
        elif type(sheet_name) == str:
            sheet_name = [sheet_name]
        for sheet_name_item in sheet_name:
            try:
                ws1 = wb1[sheet_name_item]
                ws2 = wb2[sheet_name_item]
                No_row = ws1.max_row + offset
                H1_obj = ws1[get_column_letter(1) + str(No_row - 1)]
                H1_obj.value = H1
                H1_obj.font = Font(size=15, bold=True, color='95B3DF')
                for tr in ws2:
                    No_col = 1
                    for td in tr:
                        ws1[get_column_letter(No_col) + str(No_row)].value = td.value
                        ws1[get_column_letter(No_col) + str(No_row)].fill = copy(td.fill)
                        ws1[get_column_letter(No_col) + str(No_row)].alignment = copy(td.alignment)
                        ws1[get_column_letter(No_col) + str(No_row)].font = copy(td.font)
                        ws1[get_column_letter(No_col) + str(No_row)].border = copy(td.border)
                        ws1[get_column_letter(No_col) + str(No_row)].alignment = copy(td.alignment)
                        No_col = No_col + 1
                    No_row = No_row + 1
            except:
                pass
        wb1.save(filename1)

#appender(filename1='1.xlsx',filename2 = '2.xlsx',H1='header',sheet_name='CHG0000372' )
#appender(filename1='1.xlsx',filename2 = '2.xlsx',H1='header',sheet_name='CHG0000372' )
#appender(filename1='dsad.xlsx',filename2 = 'dsad1.xlsx',H1='header')