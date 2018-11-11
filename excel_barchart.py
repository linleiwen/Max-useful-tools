import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.text import Font as draw_Font
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RegularTextRun#,RichTextProperties
from openpyxl.chart.text import RichText,RichTextProperties
from openpyxl.chart.label import DataLabelList

import re
from copy import copy


class barchart_creater():
    def __init__(self,
                 filename,
                 sheet_name= [],
                 chart_title = [],
                 xaxis_title= [],
                 yaxis_title = [],
                 stack= [],
                 chart_position="F10",
                 whiten_all_area=True,
                 font_name='Calibri',
                 axis_title_font_size=10,
                 output_filename = None
                 ):
        wb = load_workbook(filename)
        counts = len(sheet_name)
        wsDict = {}
        for count in range(counts):
            
            wsDict['ws'+str(count)] = wb[sheet_name[count]]
            
            chart = BarChart()
            chart.style = 10
            chart.type = "col"
            if stack[count] == True:
                chart.grouping = "stacked"
                chart.overlap = 100
            chart.title = chart_title[count]
            chart.y_axis.title = xaxis_title[count]
            chart.x_axis.title = yaxis_title[count]

            data = Reference(wsDict['ws'+str(count)], min_col=wsDict['ws'+str(count)].min_column + 1, min_row=wsDict['ws'+str(count)].min_row, max_row=wsDict['ws'+str(count)].max_row,
                             max_col=wsDict['ws'+str(count)].max_column)
            cats = Reference(wsDict['ws'+str(count)], min_col=wsDict['ws'+str(count)].min_column, min_row=wsDict['ws'+str(count)].min_row + 1, max_row=wsDict['ws'+str(count)].max_row,
                             max_col=wsDict['ws'+str(count)].min_column)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ## set axis title property??

            font_test = draw_Font(typeface=font_name)
            cp = CharacterProperties(latin=font_test, sz=axis_title_font_size * 100)
            chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            xPara = [Paragraph(pPr=ParagraphProperties(defRPr=cp), r=RegularTextRun(t=s)) for s in
                     xaxis_title[count].split("\n")]
            yPara = [Paragraph(pPr=ParagraphProperties(defRPr=cp), r=RegularTextRun(t=s)) for s in
                     yaxis_title[count].split("\n")]

            chart.x_axis.title.tx.rich.paragraphs = xPara
            chart.y_axis.title.tx.rich.paragraphs = yPara

            ## hide legend

            if not stack[count]:
                chart.legend = None

            wsDict['ws'+str(count)].add_chart(chart, chart_position)

            ## whiten data

            if whiten_all_area:
                table = wsDict['ws'+str(count)]["A1":"AZ200"]
                color = 'FFFFFFFF'
                font = Font(bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color=color)
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for tr in table:
                    for td in tr:
                        td.font = font
                        td.border = None
                        td.fill = fill

            ### data label

            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True


        if output_filename is None:
            wb.save(filename=filename)
        else:
            wb.save(filename=output_filename)
#
# barchart_creater(filename = 'temp0.xlsx',sheet_name=["Issue Count by Due Date","Issue Count By Owner","MTM Issue Comparison"],
#                  chart_title=["Number of Issues By Due Date","Number of outstanding Issues by Owner","Number of Issues comparison month to month by Source"],
#                  xaxis_title = ["Number of Issues",'Issue Owner','Month'],
#                  yaxis_title=['Number of Days till Due','Number of Issues','Number Of Issues'],stack = [False,False,True])