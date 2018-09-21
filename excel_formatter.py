import openpyxl
from openpyxl import load_workbook
# https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
from copy import copy


class formatter():
    def __init__(self,
                 filename,
                 sheet_name= None,
                 font_name = "Calibri",
                 font_size = 10,
                 header_font_size = 14,
                 header_bold = True,
                 froze_first_row=True,
                 add_fliter=True,
                 header_height=25,
                 add_border=True,
                 alignment_center=True,
                 specific_alignment_left = [],
                 auto_fit_column_width=True,
                 wrap_width=[],
                 fill_color_header='95B3DF',
                 hide_columns=[],
                 column_width= {},
                 output_filename=None,
                 home_page_bottom= None,
                 whiten_nontable_area= False,
                 zoom_level = 85

                 ):
        # load excel file and focus on one sheet
        if column_width is None:
            column_width = {}
        wb = load_workbook(filename)
        if sheet_name is None:
            wb.active = 0
            ws = wb.active
        else:
            ws = wb[sheet_name]
        filename_new = re.match(".*\.",filename).group()[:-1] + "_formatted.xlsx"
        # defind table, header, body
        table = ws[ws.min_row:ws.max_row]
        header = ws[ws.min_row]
        body = ws[ws.min_row + 1:ws.max_row]

        #font and font size
        font = Font(name=font_name,
                    size=font_size,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        for tr in table:
            for td in tr:
                td.font = font

        font = Font(name=font_name,
                    size=header_font_size,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        #font for header
        for td in header:
            td.font = font



        if header_bold:
            for td in header:
                font = copy(td.font)
                font.bold = True
                td.font = font


        ## frozen first row
        if froze_first_row:
            freeze_position = ws['A2']
            ws.freeze_panes = freeze_position
        ## add filter on header
        if add_fliter:
            ws.auto_filter.ref = ws.dimensions

        ## set header height
        header_obj = ws.row_dimensions[1]
        header_obj.height = header_height

        ## add_border
        if add_border:
            border = Border(left=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            diagonal=Side(border_style=None, color='FF000000'),
                            diagonal_direction=0,
                            outline=Side(border_style=None, color='FF000000'),
                            vertical=Side(border_style=None, color='FF000000'),
                            horizontal=Side(border_style=None, color='FF000000'))
            for tr in table:
                for td in tr:
                    td.border = border

        ## alignment_center
        if alignment_center:
            alignment = Alignment(horizontal='center',
                                  vertical='center',
                                  text_rotation=0,
                                  wrap_text=False,
                                  shrink_to_fit=False,
                                  indent=0)
            for tr in table:
                for td in tr:
                    td.alignment = alignment
        ## for those who have left top alignment
        for col in specific_alignment_left:
            column = ws[col]
            for td in column:
                alignment = copy(td.alignment)
                alignment.horizontal = 'left'
                td.alignment = alignment
        ###header do not change
        if alignment_center:
            for cell in header:
                alignment = copy(cell.alignment)
                alignment.horizontal = 'center'
                alignment.vertical = 'center'
                cell.alignment = alignment

        ## auto fit column width
        if auto_fit_column_width:
            column_widths = []
            for i, cell in enumerate(header):
                try:
                    column_widths[i] = len(cell.value)
                except IndexError:
                    column_widths.append(len(cell.value))
            for i, width in enumerate(column_widths):
                ws.column_dimensions[get_column_letter(i + 1)].width = width * 1.3 + 5

        ##choose column width (specific)
        for col in column_width:
            ws.column_dimensions[col].width = column_width[col]

        ## wrap
        for col in wrap_width:
            column = ws[col]
            for td in column:
                alignment = copy(td.alignment)
                alignment.horizontal = 'left'
                alignment.vertical = 'top'
                alignment.wrap_text = True
                td.alignment = alignment
                #ws.column_dimensions[col].width = 55
            ## header do not change
            alignment.horizontal ='center'
            alignment.vertical = 'center'
            ws[col + "1"].alignment=alignment

        ##fill header color
        if fill_color_header is not None:
            fill = PatternFill(start_color=fill_color_header, end_color=fill_color_header, fill_type="solid")
            for cell in header:
                cell.fill = fill

        ##hide columns
        for col in hide_columns:
            ws.column_dimensions[col].hidden = True

        if home_page_bottom is not None:
            offset = 4
            No_row = ws.max_row + offset
            Home_Bottom = ws[get_column_letter(1) + str(No_row - 1)]
            Home_Bottom.value = f'=HYPERLINK("#{home_page_bottom}!A1","Home Page")'
            Home_Bottom.font = Font(size=15, bold=True, color='95B3DF')

        ##whiten_nontable_area
        if whiten_nontable_area:
            ### define area to whiten
            whiten_col = ws.max_column + 2
            whiten_row = ws.max_row + 2
            whiten_area1 = ws[get_column_letter(whiten_col) + str(1):"AJ200"]
            whiten_area2 = ws['A' + str(whiten_row): get_column_letter(whiten_col - 1) + str(200)]
            whiten_area_col = ws[get_column_letter(whiten_col - 1) + str(1):get_column_letter(whiten_col - 1) + str(200)]
            whiten_area_row = ws['A' + str(whiten_row - 1): get_column_letter(whiten_col - 1) + str(whiten_row - 1)]

            color = 'FFFFFFFF'
            font = Font(bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color=color)
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for tr in whiten_area1:
                for td in tr:
                    td.font = font
                    td.border = None
                    td.fill = fill
            for tr in whiten_area2:
                for td in tr:
                    td.font = font
                    td.border = None
                    td.fill = fill
            for tr in whiten_area_col:
                for td in tr:
                    td.font = font
                    # td.border = None
                    td.fill = fill
            for tr in whiten_area_row:
                for td in tr:
                    td.font = font
                    # td.border = None
                    td.fill = fill

                    ## adjust zoom level
        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = zoom_level

        if output_filename is None:
            wb.save(filename=filename)
        else:
            wb.save(filename=output_filename)

#formatter('Outstanding Findings as of SEP_max_9.14.xlsx', sheet_name='raw_data', wrap_width=['J'],column_width={'J':50},header_height=36,output_filename='dsad.xlsx',home_page_bottom='Summary',whiten_nontable_area=True   )
#formatter(filename="test.xlsx", hide_columns=["C","D","F","D","S","U","T","W"])
#excel_file_name = 'test.xlsx'
#formatter(filename=excel_file_name)